# -*- coding: utf-8 -*-
# https://blog.csdn.net/weixin_55154866/article/details/128476763

import os
from cnocr import CnOcr
from openpyxl import workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, GradientFill
import time
import pdfToPng


# 设置Execl的列样式
def setCellStyle(cell):
    # 居中
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # 边框
    cell.border = Border(top=Side(style="thin", color="000000"),
                         bottom=Side(style="thin", color="000000"),
                         left=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="000000"))
    # 字体
    cell.font = Font(name="微软雅黑")
    # 背景色
    cell.fill = PatternFill("solid", fgColor="1E90FF")


# 给sheet 添加标题
def createTitle(sheet):
    # 第一列：发票代码
    cell1 = sheet.cell(1, 1)
    cell1.value = '发票代码'
    setCellStyle(cell1)

    # 第二列：发票号码
    cell2 = sheet.cell(1, 2)
    cell2.value = '发票号码'
    setCellStyle(cell2)
    # 第三列：开票日期
    cell3 = sheet.cell(1, 3)
    cell3.value = '开票日期'
    setCellStyle(cell3)
    # 第四列：购买方名称
    cell4 = sheet.cell(1, 4)
    cell4.value = '购买方名称'
    setCellStyle(cell4)
    # 第五列：购买方 纳税人识别号
    cell5 = sheet.cell(1, 5)
    cell5.value = '纳税人识别号'
    setCellStyle(cell5)
    # 第六列：税前合计
    cell6 = sheet.cell(1, 6)
    cell6.value = '税前合计'
    setCellStyle(cell6)
    # 第七列：税额
    cell7 = sheet.cell(1, 7)
    cell7.value = '税额'
    setCellStyle(cell7)
    # 第八列：税后合计
    cell8 = sheet.cell(1, 8)
    cell8.value = '税后合计'
    setCellStyle(cell8)
    # 第九列：备注
    cell9 = sheet.cell(1, 9)
    cell9.value = '备注'
    setCellStyle(cell9)


op_flag = True
ocr = CnOcr()
while op_flag:
    try:
        print(f"请输入需要分析的发票图片路径====》")
        filePath = input()
        op_flag = False
        # 路径不存在
        if not os.path.exists(filePath):
            print(f"输入路径不存在！路径为：${filePath}")
        else:
            # 读取文件夹中的所有文件
            fileList = os.listdir(filePath)
            # 先把PDF转成图片
            for pafFile in fileList:
                if (pafFile.endswith(".pdf") or pafFile.endswith(".PDF") or
                        pafFile.endswith(".ofd") or pafFile.endswith(".OFD")):
                    pdf_path = filePath + '\\' + pafFile
                    print(f"路径：{filePath}")
                    pdfToPng.pdf_image(pdf_path, filePath, 2, 2)

            # 创建execl且默认会创建一个sheet（名称为Sheel)
            wb = workbook.Workbook()
            sheet = wb.worksheets[0]
            # 设置标题
            createTitle(sheet)
            # 读取文件夹中的所有文件
            imageList = os.listdir(filePath)
            # 第二行开始写数据
            row_index = 2
            # 过滤：只保留png结尾的图片
            for img in imageList:
                if (img.endswith(".png") or img.endswith(".PNG") or
                        img.endswith(".jpg") or img.endswith(".JPG") or
                        img.endswith(".jpeg") or img.endswith(".JPEG")):
                    print(f"处理图片：{img}")

                    res = ocr.ocr(filePath + '\\' + img)
                    cell_index = 0
                    name_count = 0
                    sheet.row_dimensions[row_index].height = 20
                    for each in res:
                        cell_index = cell_index + 1
                        cell_val = str(each['text']).strip()
                        # 第一列：发票代码
                        if cell_val.startswith('发票代码:'):
                            str_val = []
                            if cell_val.find('|') >= 0:
                                str_val = cell_val.split('|')
                            else:
                                str_val = cell_val.split(':')
                            cell1 = sheet.cell(row_index, 1)
                            if len(str_val) > 1:
                                cell1.value = str_val[1]
                            else:
                                cell1.value = str_val[0]
                        # 第二列：发票号码
                        if cell_val.startswith('发票号码：'):
                            next_val = res[cell_index]
                            cell2 = sheet.cell(row_index, 2)
                            cell2.value = next_val['text']
                        if cell_val.startswith('发票号码:'):
                            str_val = cell_val.split(':')
                            cell2 = sheet.cell(row_index, 2)
                            cell2.value = str_val[1]
                        # 第三列：开票日期
                        if cell_val.startswith('开票日期:'):
                            str_val = cell_val.split(':')
                            cell3 = sheet.cell(row_index, 3)
                            cell3.value = str_val[1]
                        if cell_val.startswith('开票日期：'):
                            str_val = cell_val.split('：')
                            cell3 = sheet.cell(row_index, 3)
                            cell3.value = str_val[1]
                        # 第四列：购买方 名称
                        if cell_val.startswith('称：'):
                            name_count = name_count + 1
                            if name_count < 2:
                                str_val = cell_val.split('：')
                                cell4 = sheet.cell(row_index, 4)
                                cell4.value = str_val[1]
                        # 第五列：购买方 纳税人识别号
                        if cell_val.startswith('纳税人识别号:'):
                            str_val = cell_val.split(':')
                            cell5 = sheet.cell(row_index, 5)
                            cell5.value = str(str_val[1]).strip()
                        # 第六列：税前合计
                        if cell_val.startswith('税额'):
                            other_val = str(res[cell_index + 4]['text']).strip()
                            if not other_val.startswith('|13%') and not other_val.startswith(
                                    '|3%') and not other_val.startswith('|1%'):
                                # 税后金额
                                temp_val = other_val.split('|')
                                get_money = str(temp_val[0]).strip()
                                cell8 = sheet.cell(row_index, 8)
                                cell8.value = get_money
                                # 税费
                                sui_val = str(res[cell_index + 6]['text']).strip()
                                cell7 = sheet.cell(row_index, 7)
                                cell7.value = sui_val
                                # 税前金额
                                all_money = str(res[cell_index + 13]['text']).strip()
                                list_all_money = all_money.split('Y')
                                cell6 = sheet.cell(row_index, 6)
                                if len(list_all_money) > 1:
                                    cell6.value = list_all_money[1].strip()
                                else:
                                    cell6.value = list_all_money[0].strip()
                            else:
                                # 税后金额
                                get_money = str(res[cell_index + 8]['text']).strip()
                                list_get_money = get_money.split('¥')
                                cell8 = sheet.cell(row_index, 8)
                                cell8.value = list_get_money[1].strip()
                                # 税费
                                sui_val = str(res[cell_index + 9]['text']).strip()
                                list_sui_val = sui_val.split('¥')
                                cell7 = sheet.cell(row_index, 7)
                                cell7.value = list_sui_val[1].strip()
                                # 税前金额
                                all_money = str(res[cell_index + 12]['text']).strip()
                                list_all_money = all_money.split('X')
                                cell6 = sheet.cell(row_index, 6)
                                cell6.value = list_all_money[1].strip()
                        # 第九列：备注
                        if cell_val.startswith('备'):
                            remark_val = str(res[cell_index - 3]['text']).strip()
                            cell9 = sheet.cell(row_index, 9)
                            cell9.value = remark_val

                    row_index = row_index + 1

            # 设置高度和宽度
            sheet.row_dimensions[1].height = 25
            sheet.column_dimensions["A"].width = 17
            sheet.column_dimensions["B"].width = 17
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 50
            sheet.column_dimensions["E"].width = 30
            sheet.column_dimensions["F"].width = 15
            sheet.column_dimensions["G"].width = 15
            sheet.column_dimensions["H"].width = 15
            sheet.column_dimensions["I"].width = 50

            time_str = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            wb.save(filePath + '\\图片信息_' + time_str + '.xls')
            print(f"处理完成，请验证！！")

        op_flag = True
    except Exception as err:
        print(f"异常了，信息为: ${err}")
